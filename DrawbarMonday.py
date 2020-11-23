from selenium import webdriver
from selenium.webdriver.common.keys import Keys  # Allows access to non character keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
import re

def locate_by_name(web_driver, name):
    """Clicks on something by name."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.NAME, name))).click()
    # Returns nothing

def locate_by_id(web_driver, id):
    """Clicks on something by id."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.ID, id))).click()
    # Returns nothing

def locate_by_class(web_driver, class_name):
    """Clicks on something by class name."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, class_name))).click()
    # Returns nothing

def PRP(driver):
    comp_name = []
    comp_qty = []
    comp_des = []
    def collect_drawbars(overall_qty):
        # Returns a list of all needed components from a Bill of Materials.
        # Used for traversing Bill of Materials since component numbers and their 
        # descriptions use the NoWrap class
        no_wraps = driver.find_elements_by_class_name("NoWrap")
        nw_row = 4  # Initialize to 4 since the boxes we need start at 4
        # Keep track of the previous box since it provides help with reading info.
        previous_link = None
        # Component numbers and their qty's are links
        material_links = driver.find_elements_by_xpath("//a[@href]")
        for m_link in material_links:
            if re.search("Plexus_Control", m_link.get_attribute("href")):
                if re.search("-P", m_link.text) or re.search("-E", m_link.text):
                    partNo = m_link.text.split("@")
                    comp_name.append(partNo[0].strip())
                    qty = re.findall("\d", previous_link.text)
                    # Multiply component qty by part qty
                    comp_qty.append(int(''.join(qty)) * totals_needed[overall_qty])
                    comp_des.append(no_wraps[nw_row].text)
                # Go to the next description after each component box
                nw_row += 2
            # This gets updated after each link
            previous_link = m_link

    # Navigate to PRP page
    action = ActionChains(driver)
    action.key_down(Keys.CONTROL).send_keys('M').key_up(Keys.CONTROL).perform()
    action = 404
    time.sleep(1)
    action = ActionChains(driver)
    action.send_keys("PRP").send_keys(Keys.RETURN).perform()
    action = 404
    time.sleep(1)
    action = ActionChains(driver)
    action.send_keys(Keys.RETURN).perform()

    # Fill out the search criteria
    time.sleep(2)
    input_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "fltPRPWindow")))
    input_box.clear()  # Clearing this box will automatically leave a 1 by default
    locate_by_id(driver, "lblRequirementsOnly")
    locate_by_id(driver, "lblSuppressForecast")
    input_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "flttxtPRPPlanningGroupKey")))
    input_box.send_keys("Drawbar Planning")
    locate_by_id(driver, "btnSearch_Label")

    # Start scrapping the results
    time.sleep(3)
    list_size = 0  # Amount of parts that came up
    totals_needed = []  # Qty needed for each part
    links = driver.find_elements_by_xpath("//a[@href]")  # Get every link
    for link in links:
        # Find the total amount of parts
        if re.search("Plexus_Control", link.get_attribute("href")):
            list_size += 1
        # Get the qty for each part
        elif re.search("Job_Form", link.get_attribute("href")):
            totals_needed.append(int(link.text) * -1)  # There should be two for each part
    del totals_needed[0::2]  # But only the second box of each part is needed
    # len(totals_needed) should equal list_size
    if list_size < 1:  # This should never happen, but just in case
        raise Exception("There's nothing to scrape")
    time.sleep(2)

    for index in range(list_size):
        # Links on a page change if you navigate to
        # another page and come back. Thus, I can't
        # find all relevant links and store them into
        # a list; doing so would make all of them
        # moot if I click on a part.
        elems = driver.find_elements_by_xpath("//a[@href]")
        encountered = 0  # Keep track of how many relevant links have been found
        for elem in elems:
            if re.search("Plexus_Control", elem.get_attribute("href")):
                # Search for part links
                if encountered == index: 
                    elem.click()  # Click on the part
                    time.sleep(1)
                    submenu_cells = driver.find_elements_by_class_name("CellBottom")
                    submenu_cells[3].click()  # Click on Bill of Materials
                    time.sleep(1)
                    collect_drawbars(index)  # Scrap information
                    locate_by_class(driver, "left-arrow-purple")  # Click the back button
                    time.sleep(1)
                    locate_by_class(driver, "left-arrow-purple")
                    time.sleep(1)
                    break  # Exit the nearest for loop
                encountered += 1

    # Combine duplicates and put everything scrapped into another data structure
    seen = {}
    for index, name in enumerate(comp_name):
        if name not in seen:
            seen[name] = [comp_qty[index], comp_des[index]]
        else:
            seen[name][0] += comp_qty[index]
    del comp_name
    del comp_qty
    del comp_des
    del totals_needed
    time.sleep(1)

    # Write data to new excel sheet
    wb_obj = openpyxl.Workbook()
    sheet_obj = wb_obj.active
    headers = ["Part No", "Description", "Qty", "Location", "Deliver to", "Status", "Request From"]
    for i in range(2, 9):
        sheet_obj.cell(row=1, column=i).value = headers[i-2]
    for index, key in enumerate(seen):
        sheet_obj.cell(row=index+2, column=1).value = "Shawn"
        sheet_obj.cell(row=index+2, column=2).value = key
        sheet_obj.cell(row=index+2, column=3).value = seen[key][1]
        sheet_obj.cell(row=index+2, column=4).value = seen[key][0]
        sheet_obj.cell(row=index+2, column=5).value = "ADB01"
        sheet_obj.cell(row=index+2, column=6).value = "Paint"
        sheet_obj.cell(row=index+2, column=7).value = "REQUESTED"
        sheet_obj.cell(row=index+2, column=8).value = "ST Pull"
    # This creates a new excel workbook in the program's directory
    wb_obj.save("Drawbar request list.xlsx")

try:
    # Getting into Plex
    driver = webdriver.Chrome("chromedriver.exe")
    driver.get("https://accounts.plex.com/interaction/fea73869-0eda-4f67-b381-c167be521da6#ilp=woW7Rk4HS5ijknMk0L8Jjl8&ie=1606149525001")
    parent = "//form[@class='form-horizontal']//div[@class='plex-idp-wrapper']"  # Allows access to input fields, which are hidden
    # Enter in company code
    form = driver.find_element_by_xpath(parent + "//div[@id='companyCodeInput']//div[@class='col-sm-12']//input[@id='inputCompanyCode3']")
    form.send_keys("wanco")
    action = ActionChains(driver)
    action.send_keys(Keys.RETURN).perform()
    time.sleep(.5)
    # Enter in username
    form = driver.find_element_by_xpath(parent + "//div[@id='usernameInput']//div[@class='col-sm-12']//input[@id='inputUsername3']")
    form.send_keys("w.mc.tester")
    action.perform()
    time.sleep(.5)
    # Enter in password
    form = driver.find_element_by_xpath(parent + "//div[@id='passwordInput']//div[@class='col-sm-12']//input[@id='inputPassword3']")
    form.send_keys("test1wanco")
    action.perform()
    time.sleep(.5)
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(3)
    PRP(driver)
except Exception as e:
    print("An error was encountered:")
    print(e)
finally:
    driver.quit()
